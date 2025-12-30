#!/usr/bin/env python3
"""
Script to parse T4T Year End Feedback Excel file and extract all survey data
"""
import sys
import json
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip3 install openpyxl")
    print("Or use: python3 -m pip install --user openpyxl")
    sys.exit(1)

def parse_excel(file_path):
    """Parse the Excel file and extract all survey responses"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Get headers (first row)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else "")
    
    # Get all data rows
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip completely empty rows
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value
            data.append(row_dict)
    
    return headers, data

def analyze_data(headers, data):
    """Analyze the data and create statistics"""
    total_responses = len(data)
    
    results = {
        'totalResponses': total_responses,
        'columns': headers,
        'data': {}
    }
    
    # Analyze each column
    for header in headers:
        if not header:
            continue
        
        # Count responses for this column
        counter = Counter()
        multi_select_counter = Counter()
        
        for row in data:
            value = row.get(header, '')
            if value:
                # Handle multi-select (comma or semicolon separated)
                if isinstance(value, str) and (',' in value or ';' in value):
                    # Split by comma or semicolon
                    parts = [p.strip() for p in value.replace(';', ',').split(',')]
                    for part in parts:
                        if part:
                            multi_select_counter[part] += 1
                else:
                    counter[str(value).strip()] += 1
        
        # Use multi-select if it has data, otherwise use single select
        if multi_select_counter:
            results['data'][header] = dict(multi_select_counter)
        elif counter:
            results['data'][header] = dict(counter)
    
    return results

def main():
    file_path = '/Users/quinn/Downloads/T4T Year End Feedback (Responses) (1).xlsx'
    
    print(f"Reading Excel file: {file_path}")
    headers, data = parse_excel(file_path)
    
    print(f"\nFound {len(headers)} columns")
    print(f"Found {len(data)} responses")
    
    print("\nColumn names:")
    for i, header in enumerate(headers):
        if header:
            print(f"  {i}: {header}")
    
    print("\nAnalyzing data...")
    results = analyze_data(headers, data)
    
    # Print summary for each question
    print("\n" + "="*80)
    print("DATA SUMMARY")
    print("="*80)
    
    for header, counts in results['data'].items():
        print(f"\n{header}:")
        sorted_counts = sorted(counts.items(), key=lambda x: -x[1])
        for key, value in sorted_counts:
            print(f"  {key}: {value}")
        print(f"  Total: {sum(counts.values())}")
    
    # Save to JSON for reference
    output_file = 'survey_data_analysis.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    
    print(f"\n\nFull analysis saved to: {output_file}")
    print(f"\nTotal responses: {results['totalResponses']}")

if __name__ == '__main__':
    main()

